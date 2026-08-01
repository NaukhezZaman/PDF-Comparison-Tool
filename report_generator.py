from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

class ReportGenerator:

    def __init__(self):
        pass

    def generate_report(self, output_file, source_file, target_file, differences):

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        difference_sheet = workbook.create_sheet(title="Differences")

        self._populate_summary_sheet(
            summary_sheet,
            source_file,
            target_file,
            differences
        )

        self._populate_difference_sheet(
            difference_sheet,
            differences
        )

        self._apply_formatting(workbook)

        self._apply_row_colors(difference_sheet)

        workbook.save(output_file)

        print(f"Excel report generated successfully: {output_file}")

    def _populate_summary_sheet(self, summary_sheet, source_file, target_file, differences):
        generated_on = datetime.now().strftime("%d-%b-%Y %I:%M %p")
        modified = 0
        deleted = 0
        inserted = 0

        for difference in differences:

            if difference.difference_type == "MODIFIED":
                modified += 1

            elif difference.difference_type == "DELETED":
                deleted += 1

            elif difference.difference_type == "INSERTED":
                inserted += 1

        summary_sheet.append(["Metric", "Value"])
        summary_sheet.append(["Source File", source_file])
        summary_sheet.append(["Target File", target_file])
        summary_sheet.append(["Generated On", generated_on])
        summary_sheet.append(["Total Differences", len(differences)])
        summary_sheet.append(["Modified", modified])
        summary_sheet.append(["Deleted", deleted])
        summary_sheet.append(["Inserted", inserted])

    def _populate_difference_sheet(self, difference_sheet, differences):

        difference_sheet.append([
            "Page",
            "Type",
            "Source",
            "Target",
            "Block",
            "Line",
            "Word"
        ])

        for difference in differences:
            difference_sheet.append([
                difference.page,
                difference.difference_type,
                difference.source_text,
                difference.target_text,
                difference.block,
                difference.line,
                difference.word
            ])

    def _apply_formatting(self, workbook):

        bold_font = Font(bold=True)

        for sheet in workbook.worksheets:

            sheet.freeze_panes = "A2"

            sheet.auto_filter.ref = sheet.dimensions

            for cell in sheet[1]:
                cell.font = bold_font

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            for column_cells in sheet.columns:

                max_length = 0

                column_letter = get_column_letter(
                    column_cells[0].column
                )

                for cell in column_cells:

                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                sheet.column_dimensions[
                    column_letter
                ].width = max_length + 3

    def _apply_row_colors(self, difference_sheet):

        modified_fill = PatternFill(
            fill_type="solid",
            start_color="FFF59D",  # Slightly richer yellow
            end_color="FFF59D"
        )

        deleted_fill = PatternFill(
            fill_type="solid",
            start_color="E57373",  # Medium red
            end_color="E57373"
        )

        inserted_fill = PatternFill(
            fill_type="solid",
            start_color="A5D6A7",  # Slightly richer green
            end_color="A5D6A7"
        )

        for row in difference_sheet.iter_rows(min_row=2):
            difference_type = row[1].value
            if difference_type == "MODIFIED":

                fill = modified_fill

            elif difference_type == "DELETED":

                fill = deleted_fill

            elif difference_type == "INSERTED":

                fill = inserted_fill

            else:
                continue

            for cell in row:
                cell.fill = fill